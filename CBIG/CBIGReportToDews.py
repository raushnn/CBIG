import traceback
import openpyxl
import pandas as pd
import paramiko
from pretty_html_table import build_table
import pysftp as pysftp
from SetLogger import Logs
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.client_context import ClientContext
import time
from datetime import date
import datetime
from SendEmail import SendEmail as mail
from CBIG.DatabaseConfigFile import *
# from DatabaseConfigFile import *
from SharepointSettings import settings_ReportToDews
import pyodbc
from pathlib import Path
from openpyxl import load_workbook
import numpy as np



class AutoReportToDews:

    # handle sql
    def managesqlconnection(self,  duns,  companyname,  logger):

        try:
            # CompanyName=regex.escape(CompanyName, special_only=False)
            sql_conn = pyodbc.connect(
                'DRIVER=' + iAccess["DRIVER"] + ';SERVER=' + iAccess["SERVER"] +
                ';DATABASE=' + iAccess["DATABASE"] + ';UID=' + iAccess["UID"] + ';PWD=' + iAccess[
                    "PWD"] + '')
            cursor = sql_conn.cursor()

            sql2 = """\
            EXEC VerifyDUNSLegalName   @DUNS = ?, @CompanyName = ?;"""

            values2 = (str(duns),  companyname)

            cursor.execute(sql2,  values2)
            rc = cursor.fetchval()
            print(rc)
            sql_conn.commit()
            cursor.close()
            sql_conn.close()
            logger.info("Data Stored : Process Completed")
            return rc
        except Exception as ex:
            logger.error("SQL Error")
            logger.error(str(ex))
            return 0

    # Constructor
    def __init__(self):
        self.DFSharepointFolderDetails = pd.DataFrame(
            columns=['Folder Name',  'File Count',  'Folder Status',  'ProcessedDateTime'])

    # Sharepoint API Integration
    def sharepointoperation(self,  action,  file,  logger,  fileType):

        SharepointFolderDetails = []
        AutoReportToDewsObj = AutoReportToDews()
        try:

            ctx = ClientContext(settings_ReportToDews.get('team_site_url')).with_credentials(
                ClientCredential(settings_ReportToDews['client_credentials']['client_id'],
                                 settings_ReportToDews['client_credentials']['client_secret']))

            if (action == "Upload File"):
                if (fileType == "LOGS"):
                    path = file
                    folder_url = '/Shared Documents/CBIG/Logs_Stats/'
                    with open(path,  'rb') as content_file:
                        file_content = content_file.read()
                    target_folder = ctx.web.get_folder_by_server_relative_url(folder_url)
                    name = os.path.basename(path)
                    target_folder.upload_file(name,  file_content)
                    ctx.execute_query()
                elif (fileType == "CSV"):
                    folder_url = '/Shared Documents/CBIG/Processed_Files/'
                    for file in os.listdir(OutPutFilesReportToDews["OutPutFilesReportToDews"]):
                        try:
                            if not file.endswith(".csv"):
                                continue
                            target_folder = ctx.web.get_folder_by_server_relative_url(folder_url)
                            path = os.path.join(OutPutFilesReportToDews["OutPutFilesReportToDews"],  file)
                            name = os.path.basename(
                                os.path.join(OutPutFilesReportToDews["OutPutFilesReportToDews"],  file))
                            with open(path,  'rb') as content_file:
                                file_content = content_file.read()

                            target_folder.upload_file(name,  file_content)
                            ctx.execute_query()
                        except Exception as ex:
                            logger.error("Could not upload the file to csv")

                return "File uploaded"

            elif (action == "Download File"):
                try:
                    folder_url = '/Shared Documents/CBIG'
                    folders = ctx.web.get_folder_by_server_relative_url(folder_url).folders
                    ctx.load(folders)
                    ctx.execute_query()
                    FileListCount = []
                    t0 = time.time()
                    try:
                        # len(folders)
                        for folder,  i in zip(folders,  range(0,  len(folders))):
                            FileCoun = 0
                            try:
                                FolderName = str(folder.properties["Name"])
                                IgnoreFolderNames = ['Logs_Stats', 'Processed_Files']
                                files = ctx.web.get_folder_by_server_relative_url(
                                    folder_url + FolderName).files
                                if FolderName in IgnoreFolderNames:
                                    continue
                                # files = folder.files
                                ctx.load(files)
                                ctx.execute_query()
                                if (len(files) > 0):
                                    List = [FolderName,  len(files),  "File Available",  str(today) + "_" + str(logtime)]
                                else:
                                    List = [FolderName,  len(files),  "Folder Empty",  str(today) + "_" + str(logtime)]
                                SharepointFolderDetails.append(List)

                                AutoReportToDewsObj.DFSharepointFolderDetails.loc[
                                    len(AutoReportToDewsObj.DFSharepointFolderDetails)] = List
                                for cur_file,  i in zip(files,  range(0,  len(files))):
                                    # if 'ReportToDews' in (str(cur_file.properties["Name"])):
                                    FileCoun = FileCoun + 1

                                    download_FileName = os.path.join(InPutFilesReportToDews[
                                                                         "InPutFilesReportToDews"],
                                                                     FolderName + "_" + os.path.basename(
                                                                         cur_file.properties["Name"]))
                                    file_url = '/sites/ReportToDews/Shared Documents/CBIG/' + FolderName+"/" + os.path.basename(
                                        cur_file.properties["Name"])

                                    with open(download_FileName,  "wb") as local_file:
                                        file = ctx.web.get_file_by_server_relative_url(file_url).download(
                                            local_file).execute_query()

                                    logger.info("File name: {0}".format(
                                        str(FolderName + "_" + cur_file.properties["Name"])))
                                    # delete the file from sharepoint
                                    # Below code To be enabled during production
                                    file.delete_object()
                                    ctx.execute_query()
                                FileListCount.append(FileCoun)
                            except Exception as ex:
                                logger.error(str(ex) + traceback.format_exc())
                                continue

                    except Exception as ex:
                        logger.error(str(ex) + traceback.format_exc())

                    if folders:
                        t1 = time.time()
                        Series = pd.Series(FileListCount)
                        AutoReportToDewsObj.DFSharepointFolderDetails["FileCount"] = Series
                        elapsed = time.strftime("%H:%M:%S %Z",  time.gmtime(t1 - t0))
                        logger.info("Files has been downloaded")
                        logger.info("Total time to download the files is " + str(elapsed))
                        return "Folder has been downloaded",  AutoReportToDewsObj
                    else:
                        logger.info("No folders to download")
                        return "Folder has not been downloaded",  AutoReportToDewsObj

                except Exception as ex:
                    logger.error(str(ex) + traceback.format_exc())
                    logger.error("SharePoint Folder has not been downloaded")
                    return "Folder has not been downloaded",  AutoReportToDewsObj

            elif (action == "Delete File"):
                try:
                    folder_url = '/Shared Documents/CBIG/Processed_Files/'
                    files = ctx.web.get_folder_by_server_relative_url(folder_url)

                    ctx.load(files)
                    ctx.execute_query()
                    FileListCount = []
                    # len(folders)
                    for file,  i in zip(files,  range(0,  len(files))):
                        try:
                            TimeCreated = (str(file.properties['TimeCreated']).split('T')[0])
                            TimeCreated = datetime.strptime(TimeCreated,  '%Y-%m-%d')
                            YesterdayDate = datetime.strptime(Yesterday["Yesterday"],  '%Y-%m-%d')
                            todaydate = datetime.strptime(Today["Today"],  '%Y-%m-%d')

                            duration = (todaydate - TimeCreated).days
                            # Change the value from 7 to 30 days in production
                            if duration>7:
                                file.delete_object()
                                ctx.execute_query()
                                logger.info("CSV File was deleted")
                            # else:
                            #     logger.info("No Files found older than 7 days")
                            # FileListCount.append(FileCoun)
                        except Exception as ex:
                            logger.error(str(ex) + traceback.format_exc())
                            continue

                    return "CSV Files have been deleted",  AutoReportToDewsObj

                except Exception as ex:
                    logger.error(str(ex) + traceback.format_exc())

            elif (action == "Delete Log File"):
                try:
                    folder_url = '/Shared Documents/SupremeCourt/AutoFin-NonFinUpload/Logs_Stats/NonFinancialUploadLogs/'
                    files = ctx.web.get_folder_by_server_relative_url(folder_url).files

                    ctx.load(files)
                    ctx.execute_query()

                    # len(folders)
                    for file,  i in zip(files,  range(0,  len(files))):
                        try:
                            timecreated = (str(file.properties['TimeCreated']).split('T')[0])
                            timecreated = datetime.strptime(timecreated,  '%Y-%m-%d')
                            sevendaysold = datetime.strptime(SevenDaysOldFile["SevenDaysOldFile"],  '%Y-%m-%d')
                            todaydate = datetime.strptime(Today["Today"],  '%Y-%m-%d')
                            if timecreated < sevendaysold:
                                file.delete_object()
                                ctx.execute_query()
                                logger.info("Log files older than 7 days were deleted")
                            # FileListCount.append(FileCoun)
                        except Exception as ex:
                            logger.error(str(ex) + traceback.format_exc())
                            continue

                    return "Log Files have been deleted",  AutoReportToDewsObj

                except Exception as ex:
                    logger.error(str(ex) + traceback.format_exc())

        except Exception as ex:
            logger.error(str(ex) + traceback.format_exc())

    # Update the section and field ID
    def UpdateFieldIDDataframes(self,  df,  logger):
        try:
            sheetname = str(df["SheetName"])
            if "Mgmt" in  sheetname:
                df['sid'] = df.groupby('duns_no').cumcount() + 1
                df['sid'] = -df['sid'].astype(int)

            elif "Bank Loop_Non Loop" in sheetname:
                #  Loops

                df['id'] =df.assign(temp=~df.duplicated(subset=['duns_no', 'Indian Bank Name_Bank:bb'])).groupby('duns_no')[
                    'temp'].cumsum()

                df['id'] = -df['id'].astype(int)
                # FIELDS Loops
                df['Fieldid'] = df.groupby(['duns_no','Indian Bank Name_Bank:bb']).cumcount() + 1
                df['Fieldid'] = -df['Fieldid'].astype(int)

            else:
                df['id'] = df.groupby('duns_no').cumcount() + 1
                df['id'] = -df['id'].astype(int)
                df.applymap(lambda x: x.strip() if type(x) == str else x)
                if 'SIC' in sheetname:
                    df['SIC extension_General:ce'] = df['SIC extension_General:ce'].astype(str).apply(lambda x: x.zfill(4))

        except Exception as ex:
            logger.error(str(ex))

    # Data Type conversion
    def validate_data(self,df,dffielddict,logger):
        try:
            sheetname = str(df["SheetName"])
            if "Mgmt" in sheetname:
                df = df.astype(dffielddict)
                df['Identity Number_Mgmtbac:EL'] = df['Identity Number_Mgmtbac:EL'].apply(lambda x: '{0:0>8}'.format(x))
            else:
                df = df.astype(dffielddict,
                     errors='ignore')
            return df
        except Exception as ex:
            logger.error(str(ex))

    # Process the Input Files
    def autofinnonfin(self,  logger):
        global StatsList, ValidationStatsList
        FTPStats = []
        StatsList, ValidationStatsList = [], []
    
        Inputdirectory = InPutFilesReportToDews["InPutFilesReportToDews"]
        if not os.path.exists(Inputdirectory):
            os.makedirs (Inputdirectory)
        # if os.listdir(Inputdirectory) != []:
        #         print(Inputdirectory)
        #         print("Report to Dews folder is not empty")
        # else :
        #         print("Report to Dews folder is empty")
        #         return
        #
            
            
            # if Inputdirectory:
            #     print(Inputdirectory)
            #     print("Folder is not empty")
            # elif not os.listdir():
            #     print(Inputdirectory)
            #     print("folder is empty")
       
        Outputdirectory = OutPutFilesReportToDews["OutPutFilesReportToDews"]
        for i in os.listdir (Outputdirectory):
            if '.csv' in i:
                os.remove(Outputdirectory+'\\'+i)
        df_FinNonFin=pd.DataFrame()
        dataAll = [
        ['General non loop',  'General non loop',  'NL', 'N', 'A:Z', ['DUNS_General:aa', 'Name_General:ab', 'Report base date_General:zd', 'Report type_General:zh', 'Street Name 1_General:da', 'Street Name 2_General:db', 'State_General:di', 'India-cities_General:im', 'Country_General:dd', 'Post code_General:dg', 'WWW address_General:aw', 'Email address_General:ax', 'LOB_General:ca', 'Total employees_General:cr', 'Local_General:et', '%_General:eu', 'International_General:ev', '%_General:ew', 'Local purchase terms_General:fe', 'Import terms_General:fj', 'Local_General:du', '%_General:dv', 'International_General:dw', '%_General:dx', 'Local sales terms_General:ee', 'Export sales terms_General:ej', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_Insurance',  'Gen_Insurance',  'NL', 'Y', 'A:C', ['DUNS_General:aa', 'Name_General:ab', 'Insurance_General:in', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_Award',  'Gen_Award', 'L', 'Y', 'A:G', ['DUNS_General:aa', 'Name_General:ab', 'Type of award_General:TA', 'Category_General:AC', 'Awardees_General:AE', 'Awarder_General:AW', 'Year1_General:AY', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_Membership',  'Gen_Membership', 'L', 'Y', 'A:F', ['DUNS_General:aa', 'Name_General:ab', 'Membership details_General:HP', 'Membership number_General:HQ', 'Valid from_General:HR', 'Valid till_General:HS', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_Registration',  'Gen_Registration', 'L', 'Y', 'A:D', ['DUNS_General:aa', 'Name_General:ab', 'Registration type_General:RT', 'Registration Number_General:RN', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_TradeStyle', 'Gen_TradeStyle', 'L', 'Y', 'A:D', ['DUNS_General:aa', 'Name_General:ab', 'Tradestyle_General:bd', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_ProvFins', 'Gen_ProvFins', 'L', 'Y', 'A:L', ['DUNS_General:aa', 'Name_General:ab', 'Statement end date_General:qn', 'Months covered_General:qo', 'Unit of size_General:qq', 'Currency_General:qr', 'Sales type_General:qt', 'Amount_General:qu', 'Net profit before tax type_General:qx', 'Amount_General:qy', 'Net profit after tax type_General:qz', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_ISOCert', 'Gen_ISOCert', 'L', 'Y', 'A:G', ['DUNS_General:aa', 'Name_General:ab', 'ISO_General:hr', 'Certificate no._General:TN', 'Valid from_General:VF', 'Valid to_General:VT', 'Valid to_General:IC', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_Product', 'Gen_Product', 'L', 'Y', 'A:C', ['DUNS_General:aaName_General:abProduct name_General:gb', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_CustomerTable', 'Gen_CustomerTable', 'L', 'Y', 'A:F', ['DUNS_General:aa', 'Name_General:ab', 'Major customer name_General:ep', 'Country_General:eq', '% sales_General:es', 'Length of relationship_General:PR', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_ExportTerms', 'Gen_ExportTerms', 'L', 'Y', 'A:E', ['DUNS_General:aa', 'Name_General:ab', 'Terms_General:ek', 'Net days_General:el', 'To net days_General:em', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_LocalSalesTerms', 'Gen_LocalSalesTerms', 'L', 'Y', 'A:E', ['DUNS_General:aa', 'Name_General:ab', 'Terms_General:eg', 'Net days_General:eh', 'To net days_General:ei', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_CountrywiseExport', 'Gen_CountrywiseExport', 'L', 'Y', 'A:E', ['DUNS_General:aa', 'Name_General:ab', 'Country#_General:dz', 'Region#_General:ea', '%_General:ec', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_localPurchTerms', 'Gen_localPurchTerms', 'L', 'Y', 'A:E', ['DUNS_General:aa', 'Name_General:ab', 'Terms_General:fg', 'Net days_General:fh', 'To net days_General:fi', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['gen_ImportTerms', 'gen_ImportTerms', 'L', 'Y', 'A:E', ['DUNS_General:aa', 'Name_General:ab', 'Terms_General:fk', 'Net days_General:fl', 'To net days_General:fm', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_Countrywise Purchase', 'Gen_Countrywise Purchase', 'L', 'Y', 'A:E', ['DUNS_General:aa', 'Name_General:ab', 'Country#_General:ey', 'Region#_General:ez', '%_General:fb', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['General_Source', 'General_Source', 'L', 'N', 'A:D', ['DUNS_General:aa', 'Name_General:ab', 'Source_General:DH', 'Source date_General:DI', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['General_Telephone', 'General_Telephone', 'L', 'Y', 'A:F', ['DUNS_General:aa', 'Name_General:ab', 'Country code_General:aq', 'Area code_General:ar', 'Telephone Number_General:as', 'Telephone type_General:at', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Gen_SIC', 'Gen_SIC', 'L', 'Y', 'A:D', ['DUNS_General:aa', 'Name_General:ab', 'SIC_General:cd', 'SIC extension_General:ce', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History Non loop', 'History Non loop', 'NL', 'N', 'A:AA', ['DUNS_General:aa', 'Name_General:ab', 'Update date_History:fb', 'Legal structure_History:fd', 'Registration type_History:fa', 'Registration date_History:ff', 'Registration number_History:fi', 'Tax registration number_History:PN', 'City of registration_History:fj', 'State_History:fk', 'Country_History:fl', 'Registry details provided by_History:fr', 'Ordinary shares (Amount)_History:gt', 'Par value_History:gu', 'Currency_History:gv', 'Ordinary Shares (number)_History:gw', 'Preference shares (amount_History:gx', 'Par value_History:gy', 'Currency_History:gz', 'Pref Shares (numbers)_History:ha', 'Full/ major shareholder list_History:hz', 'Shareholding as at_History:pc', 'Last AGM date_History:ju', 'Last financial end date_History:jw', 'Start year_History:kv', 'Control year_History:kw', 'Shares publicly listed_History:gd', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History Address', 'History Address', 'L', 'Y', 'A:H', ['DUNS_General:aa', 'Name_General:ab', 'Street Name 1_History:jj', 'Street Name 2_History:jk', 'State_History:tb', 'India-cities_History:jq', 'Country_History:jr', 'Post code_History:js', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_Background', 'History_Background', 'L', 'Y', 'A:G', ['DUNS_General:aa', 'Name_General:ab', 'Date of change_History:nn', 'Nature of change_History:la', 'From legal structure_History:kt', 'Name of prev. company_History:lc', 'Date of prev. company first started_History:to', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_auditor', 'History_auditor', 'L', 'Y', 'A:C', ['DUNS_General:aa', 'Name_General:ab', 'Auditor_History:jx', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_StockExchange', 'History_StockExchange', 'L', 'Y', 'A:C', ['DUNS_General:aa', 'Name_General:ab', 'Stock Exchange listed_History:ge', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_Authorize', 'History_Authorize', 'L', 'Y', 'A:D', ['DUNS_General:aa', 'Name_General:ab', 'Authorized capital amount_History:gj', 'Currency_History:gk', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_Issued', 'History_Issued', 'L', 'Y', 'A:D', ['DUNS_General:aa', 'Name_General:ab', 'Issued capital – amount_History:gm', 'Currency_History:gn', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_Paidup', 'History_Paidup', 'L', 'Y', 'A:F', ['DUNS_General:aa', 'Name_General:ab', 'Paid up type_History:hi', 'Amount_History:hj', 'Currency_History:hk', 'Capital as at_History:na', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['History_Sharehold', 'History_Sharehold', 'L', 'Y', 'A:G', ['DUNS_General:aa', 'Name_General:ab', 'Corporate holder / proprietor / Partner_History:ib', 'Ord shares held(number)_History:ik', '% held_History:im', 'Pref shares held (number)_History:io', '% held_History:iq', 'SheetName'],{"Ord shares held(number)_History:ik": pd.Int64Dtype(),'DUNS_General:aa': np.int64,'% held_History:im':float}]
        , ['Mgmt Loop_Non Loop', 'Mgmt Loop_Non Loop', 'NL', 'Y', 'A:K', ['DUNS_General:aa', 'Name_General:ab', 'Address 1_Relcon:yg', 'Address 2_Relcon:yh', 'DUNS_General:aa', 'Name_General:ab', 'Update date_Mgmtbac:EB', 'CEO_Mgmtbac:EC', 'Salutation_Mgmtbac:ED', 'First Name_Mgmtbac:EE', 'Middle Name_Mgmtbac:EF', 'Surname_Mgmtbac:EG', 'Current Title_Mgmtbac:FP', 'Identity Type_Mgmtbac:EK', 'Identity Number_Mgmtbac:EL', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Relcn Non loop', 'Relcn Non loop', 'NL', 'N', 'A:U', ['DUNS_General:aa', 'Name_General:ab', 'Update date_Relcon:CB', 'Parent or headquarter_Relcon:DC', 'DUNS_Relcon:DE', 'Company Name_Relcon:DF', 'Address 1_Relcon:DI', 'Address 2_Relcon:DJ', 'City_Relcon:DN', 'State_Relcon:DP', 'Post Code_Relcon:DQ', 'Country_Relcon:DR', '% of shares held_Relcon:DS', 'Ultimate parent DUNS_Relcon:DZ', 'Ultimate parent company Name_Relcon:EA', 'Address 1_Relcon:ED', 'Address 2_Relcon:EE', 'City_Relcon:EI', 'State_Relcon:EK', 'Post Code_Relcon:EL', 'Country_Relcon:EM', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Rel_Subsidiary', 'Rel_Subsidiary', 'L', 'Y', 'A:K', ['DUNS_General:aa', 'Name_General:ab', 'Subsidiary DUNS_Relcon:CF', 'Subsidiary company name_Relcon:CG', 'Address 1_Relcon:CJ', 'Address 2_Relcon:CK', 'City_Relcon:CO', 'State _Relcon:CQ', 'Post code_Relcon:CR', 'Country_Relcon:CS', '% of shares owned_Relcon:CW', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Rel_Affiliate', 'Rel_Affiliate', 'L', 'Y', 'A:M', ['DUNS_General:aa', 'Name_General:ab', 'Affiliate DUNS_Relcon:BG', 'This affiliate is a joint venture_Recon:BF', 'Affiliate company name_Relcon:BH', 'Address 1_Relcon:BK', 'Address 2_Relcon:BL', 'City_Relcon:BP', 'State _Relcon:BR', 'Post code_Relcon:BS', 'Country_Relcon:BT', 'Affiliate owns shares in subject_Relcon:BX', '% of shares owned_Relcon:FC', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Rel_Branch', 'Rel_Branch', 'L', 'Y', 'A:K', ['DUNS_General:aa', 'Name_General:ab', 'Address 1_Relcon:yg', 'Address 2_Relcon:yh', 'City_Relcon:yi', 'State _Relcon:yj', 'Post code_Relcon:yf', 'Country_Relcon:yk', 'Branch location type 1_Relcon:ym', 'Owned,  leased,  rented_Relcon:zf', 'Size in Sq mtrs._Relcon:zj', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Cinvest Loop', 'Cinvest Loop', 'NL', 'Y', 'A:C', ['DUNS_General:aa', 'Name_General:ab', 'Other comments_Cinevest:uy', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Bank Loop_Non Loop', 'Bank Loop_Non Loop', 'L', 'Y', 'A:J', ['DUNS_General:aa', 'Name_General:ab', 'Update date_Bank:za', 'Indian Bank Name_Bank:bb', 'Address 1_Bank:bd', 'City_Bank:bf', 'State_Bank:bg', 'Post code_Bank:bh', 'Country_Bank:bi', 'Other comments_Bank:fo', 'SheetName'],{'DUNS_General:aa': np.int64}]
        , ['Supplier looping', 'Supplier looping', 'L', 'Y', 'A:G', ['DUNS_General:aa', 'Name_General:ab', 'Report date_Supplier:la', 'Supplier name_Supplier:lg', 'Country_Supplier:li', '% of total purchases_Supplier:lx', 'Length of relationship_Supplier:ly', 'SheetName'],{'DUNS_General:aa': np.int64}]
]
        try:
            df_FinNonFin = pd.DataFrame(dataAll,  columns=['Sheets', 'SheetName',  'Looping/NonLooping',  'Deletion',  'CellRange', 'ColumnList','dtype'])
            df_FinNonFin.set_index('Sheets', inplace=True)
        except Exception as ex:
            logger.error(str(ex))


        # No Deletion - Non Looping
        Generalnonloop_All = pd.DataFrame()
        General_Source_All = pd.DataFrame()
        History_Non_loop_All = pd.DataFrame()
        Relcn_Non_loop_All = pd.DataFrame()

        # Looping
        Gen_Insurance_All = pd.DataFrame()
        Gen_Award_All = pd.DataFrame()
        Gen_Membership_All = pd.DataFrame()
        Gen_Registration_All = pd.DataFrame()
        Gen_TradeStyle_All = pd.DataFrame()
        Gen_ProvFins_All = pd.DataFrame()
        Gen_ISOCert_All = pd.DataFrame()
        Gen_Product_All = pd.DataFrame()
        Gen_CustomerTable_All = pd.DataFrame()
        Gen_ExportTerms_All = pd.DataFrame()
        Gen_LocalSalesTerms_All = pd.DataFrame()
        Gen_CountrywiseExport_All = pd.DataFrame()
        Gen_localPurchTerms_All = pd.DataFrame()
        gen_ImportTerms_All = pd.DataFrame()
        Gen_CountrywisePurchase_All = pd.DataFrame()
        General_Telephone_All = pd.DataFrame()
        Gen_SIC_All = pd.DataFrame()
        History_Address_All = pd.DataFrame()
        History_Background_All = pd.DataFrame()
        History_auditor_All = pd.DataFrame()
        History_StockExchange_All = pd.DataFrame()
        History_Authorize_All = pd.DataFrame()
        History_Issued_All = pd.DataFrame()
        History_Paidup_All = pd.DataFrame()
        History_Sharehold_All = pd.DataFrame()
        Mgmt_Loop_Non_Loop_All = pd.DataFrame()
        Rel_Subsidiary_All = pd.DataFrame()
        Rel_Affiliate_All = pd.DataFrame()
        Rel_Branch_All = pd.DataFrame()
        Bank_Loop_NonLoop_All = pd.DataFrame()
        Supplier_looping_All = pd.DataFrame()
        Cinvest_Loop_All = pd.DataFrame()
        # ==========================================================
        # Looping - Delete DataFrames
        Gen_Insurance_All_Delete = pd.DataFrame()
        Gen_Award_All_Delete = pd.DataFrame()
        Gen_Membership_All_Delete = pd.DataFrame()
        Gen_Registration_All_Delete = pd.DataFrame()
        Gen_TradeStyle_All_Delete = pd.DataFrame()
        Gen_ProvFins_All_Delete = pd.DataFrame()
        Gen_ISOCert_All_Delete = pd.DataFrame()
        Gen_Product_All_Delete = pd.DataFrame()
        Gen_CustomerTable_All_Delete = pd.DataFrame()
        Gen_ExportTerms_All_Delete = pd.DataFrame()
        Gen_LocalSalesTerms_All_Delete = pd.DataFrame()
        Gen_CountrywiseExport_All_Delete = pd.DataFrame()
        Gen_localPurchTerms_All_Delete = pd.DataFrame()
        gen_ImportTerms_All_Delete = pd.DataFrame()
        Gen_Countrywise_Purchase_All_Delete = pd.DataFrame()
        General_Telephone_All_Delete = pd.DataFrame()
        Gen_SIC_All_Delete = pd.DataFrame()
        History_Address_All_Delete = pd.DataFrame()
        History_Background_All_Delete = pd.DataFrame()
        History_auditor_All_Delete = pd.DataFrame()
        History_StockExchange_All_Delete = pd.DataFrame()
        History_Authorize_All_Delete = pd.DataFrame()
        History_Issued_All_Delete = pd.DataFrame()
        History_Paidup_All_Delete = pd.DataFrame()
        History_Sharehold_All_Delete = pd.DataFrame()
        Mgmt_Loop_Non_Loop_All_Delete = pd.DataFrame()
        Rel_Subsidiary_All_Delete = pd.DataFrame()
        Rel_Affiliate_All_Delete = pd.DataFrame()
        Rel_Branch_All_Delete = pd.DataFrame()
        Bank_Loop_Non_Loop_All_Delete = pd.DataFrame()
        Supplier_looping_All_Delete = pd.DataFrame()
        Cinvest_Loop_All_Delete = pd.DataFrame()
        try:
            logger.info("Process Started " + str(datetime.now().strftime('%d_%m_%Y %H_%M_%S')))

            Message,  AutoFinancialObj = self.sharepointoperation("Delete File",  "",  logger,  "")
            if ('CSV Files have been deleted' in Message):
                logger.info("Step 1 : CSV Files older than 7 days have been deleted")
            Message,  AutoFinancialObj = self.sharepointoperation("Download File",  "",  logger,  "")
            if os.listdir(Inputdirectory) != []:
                logger.info("Step 1 : Folder has been downloaded from sharepoint")
            else:
                print("no files to download")
                return

            # loop through the input directory and Validating the file,  if not validated then delete it
            for file in os.listdir(Inputdirectory):
                filename = Path(os.path.join(Inputdirectory,  file))
                filename1 = filename.name
                FileWithoutExtn = os.path.splitext(filename1)[0]
                Empty = True
                IntAndDigits9 = False
                DataMatchesWitUnity=0
                ErrorMessage=""
                HistoryNonLoopEmpty= True
                GeneralSourceEmpty=True
                GeneralNonLoopChecked, HistoryNonLoopChecked, GeneralSourceChecked=False, False, False
                try:
                    if not file.endswith(".xlsx"):
                        continue

                    logger.info("File In Process : " + filename1)
                    # =============================Non Looping=====================================
                    IntAndDigits9,  NotEmpty = False,  False
                    # Creating dictionary of DataFrames Dynamically and key is the sheet name and value is dataframe
                    for (index,  colname) in df_FinNonFin.iterrows():
                        try:
                            key = colname['SheetName']

                            df = pd.read_excel(filename, sheet_name=df_FinNonFin.at[index, "SheetName"],
                                               index_col=None,
                                               na_values=['NA'],
                                               usecols=df_FinNonFin.at[index, "CellRange"],
                                               converters={'DUNS_General:aa': np.int64})
                            # else:
                            # # convert_float=False)
                            #     df = pd.read_excel(filename, sheet_name=df_FinNonFin.at[index, "SheetName"],
                            #                        index_col=None,
                            #                        na_values=['NA'],
                            #                        usecols=df_FinNonFin.at[index, "CellRange"])

                            if key=="General non loop":
                                GeneralNonLoopChecked= True
                                if not df.empty:
                                    # Check if DUNS and Legal Name is not Empty
                                    Empty = ((pd.isnull(df.loc[0,  'DUNS_General:aa'])) and (pd.isnull(df.loc[0,  'Name_General:ab']))
                                             and (pd.isnull(df.loc[0, 'Report base date_General:zd']))
                                             and (pd.isnull(df.loc[0, 'Report type_General:zh'])))

                                    # Check if DUNS Number is 9 digit
                                    df.at[0,  'DUNS_General:aa'] = int(df.iloc[0,  0])
                                    # df.iloc[0,  0] = str(df.iloc[0,  0])
                                    Length = len(str(int(df.iloc[0,  0])))
                                    # print(Length[0])
                                    if Length == 9:
                                        IntAndDigits9 = True
                                    # check if Data Matches with Unity Extract
                                    DataMatchesWitUnity = self.managesqlconnection(df.loc[0,  'DUNS_General:aa'],
                                                                                   df.loc[0,  'Name_General:ab'],  logger)

                                if Empty == True:
                                    ErrorMessage =" Duns Number or Legal Name or report Base Date or Report Type missing"
                                if IntAndDigits9== False:
                                    ErrorMessage = ErrorMessage + ",  Duns Number Validation Failed"
                                if DataMatchesWitUnity== 0:
                                    ErrorMessage = ErrorMessage+ ",  Duns Number and Legal Name did not match the Unity Extract"

                            # Source General: DH
                            # Source date General: DI
                            # 'Source_General:DH',  'Source date_General:DI',
                            elif key=="General_Source":
                                GeneralSourceChecked = True
                                if not df.empty:
                                    # check if DH and DI are not Empty
                                    GeneralSourceEmpty = ((pd.isnull(df.loc[0,  'Source_General:DH'])) and (
                                        pd.isnull(df.loc[0,  'Source date_General:DI'])))
                                    # print(df.dtypes)
                                    # Check if Duns Number is Integer and 9 digit in length
                                if GeneralSourceEmpty:
                                    ErrorMessage =ErrorMessage+" , Source_General:DH or Source date_General:DI is Empty"

                            # Registration type	History:fa
                            # Start year	History:kv
                            # Control year	History:kw
                            # 'Start year_History:kv', 	'Control year_History:kw'
                            elif key=="History Non loop":
                                HistoryNonLoopChecked = True
                                if not df.empty:
                                    # check if DUNS and Legal Name is not Empty
                                    HistoryNonLoopEmpty = ((pd.isnull(df.loc[0,  'Start year_History:kv'])) and (
                                        pd.isnull(df.loc[0,  'Control year_History:kw'])))
                                    # print(df.dtypes)
                                    # Check if Duns Number is Integer and 9 digit in length
                                if HistoryNonLoopEmpty:
                                    ErrorMessage =ErrorMessage+" ,  year_History:kv or Control year_History:kw is Empty"
                            else:
                                continue

                            if (GeneralNonLoopChecked==True and GeneralSourceChecked==True and HistoryNonLoopChecked==True):

                                if (Empty == True or IntAndDigits9 == False or DataMatchesWitUnity == 0 or GeneralSourceEmpty==True or HistoryNonLoopEmpty==True):
                                    Status = ErrorMessage
                                    ValidationStatsList.append(
                                        [str(df.loc[0,  'DUNS_General:aa']),  FileWithoutExtn,  "Input File",  "Validation Failed",
                                         str(today) + "_" + str(logtime),
                                         "Validation Failed => " + Status])
                                    # s.remove(os.path.join(Inputdirectory,  file))
                                    logger.info("Validation failed")
                                    os.remove(os.path.join(Inputdirectory,  file))
                                    break

                                elif (Empty == False and IntAndDigits9 == True and DataMatchesWitUnity == 1 and GeneralSourceEmpty==False and HistoryNonLoopEmpty==False):
                                    allsheetvalid = True
                                    Status = 'Validation Successfull'
                                    ValidationStatsList.append(
                                        [str(df.loc[0,  'DUNS_General:aa']),  FileWithoutExtn,  "Input File",  Status,
                                         str(today) + "_" + str(logtime),
                                         "All Validation Successfull"])
                                    # os.remove(os.path.join(Inputdirectory,  file))
                                    logger.info("Validation Successfull")
                                    break

                        except Exception as ex:
                            Status = 'Validation Failed'
                            ValidationStatsList.append(
                                ["",  FileWithoutExtn,  str(colname['SheetName']),  Status,
                                 str(today) + "_" + str(logtime),
                                 str(ex)])
                            logger.error(filename)
                            os.remove( InPutFilesFinancial["InPutFilesFinancial"] + filename)
                            logger.error(str(ex))
                            continue

                except Exception as ex:
                    Status = 'Validation Failed'
                    ValidationStatsList.append(["",  FileWithoutExtn,  "",  Status,  str(today) + "_" + str(logtime),
                                      str(ex)])
                    logger.error(filename)
                    logger.error(str(ex))
                    os.remove(os.path.join(Inputdirectory,  file))
                    continue

            # Creating Dynamic Dictionay of dataframe
            for file in os.listdir(Inputdirectory):
                filename = Path(os.path.join(Inputdirectory,  file))
                filename1 = filename.name
                FileWithoutExtn = os.path.splitext(filename1)[0]

                try:
                    if not file.endswith(".xlsx"):
                        continue

                    # Capture Start Time
                    t0 = time.time()

                    wb = openpyxl.load_workbook(filename)
                    logger.info("File In Process : " + filename1)

                    t1 = time.time()
                    elapsed = time.strftime("%H:%M:%S %Z",  time.gmtime(t1 - t0))
                    # =============================Non Looping=====================================
                    file_dict = {}

                    try:
                        # Creating dictionary of DataFrames Dynamically and key is the sheet name and value is dataframe
                        for (index,  colname) in df_FinNonFin.iterrows():
                            try:
                                key = colname['SheetName']
                                dtypes = colname['dtype']

                                df = pd.read_excel(filename, sheet_name=df_FinNonFin.at[index, "SheetName"],
                                                   index_col=None,
                                                   na_values=['NA'],
                                                   usecols=df_FinNonFin.at[index, "CellRange"],
                                                   # converters={'DUNS_General:aa': np.int64},
                                                   dtype=dtypes)

                                if not df.empty:
                                    df['SheetName'] = colname['SheetName']
                                    file_dict[key] = df
                                else:
                                    Status = "File Processed"
                                    StatsList.append(
                                        ["", FileWithoutExtn, str(key), Status,
                                         str(today) + "_" + str(logtime),
                                         "Worksheet is Empty"])
                            except Exception as ex:
                                Status = 'File Not Processed'
                                StatsList.append(
                                    ["",  FileWithoutExtn,  str(colname['SheetName']),  Status,
                                     str(today) + "_" + str(logtime),
                                     str(ex)])
                                logger.error(filename)
                                # os.remove( InPutFilesFinancial["InPutFilesFinancial"] + filename)
                                logger.error(str(ex))
                                continue

                    except Exception as ex:
                        logger.error(str(ex))


                    # Key is the Sheet Name in the dictionary
                    for key in file_dict:
                        try:
                            # Check if Dataframe is empty
                            if not file_dict[key].empty:
                                SheetDataFrame = file_dict[key]

                                name = SheetDataFrame
                                SheetDataFrame.rename(columns={'DUNS_General:aa': 'duns_no'}, inplace=True)

                                if not name.empty:

                                    # GENERAL
                                    if (name.at[0,  "SheetName"] == 'General non loop'):
                                        Generalnonloop_All = pd.concat([Generalnonloop_All, name], axis=0)
                                    elif (name.at[0,  "SheetName"] == 'Gen_Insurance'):
                                        Gen_Insurance_All = Gen_Insurance_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_Award'):
                                        Gen_Award_All = Gen_Award_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_Membership'):
                                        Gen_Membership_All = Gen_Membership_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_Registration'):
                                        Gen_Registration_All = Gen_Registration_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_TradeStyle'):
                                        Gen_TradeStyle_All = Gen_TradeStyle_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_ProvFins'):
                                        Gen_ProvFins_All = Gen_ProvFins_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_ISOCert'):
                                        Gen_ISOCert_All = Gen_ISOCert_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_Product'):
                                        Gen_Product_All = Gen_Product_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_CustomerTable'):
                                        Gen_CustomerTable_All = Gen_CustomerTable_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_ExportTerms'):
                                        Gen_ExportTerms_All = Gen_ExportTerms_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_LocalSalesTerms'):
                                        Gen_LocalSalesTerms_All = Gen_LocalSalesTerms_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_CountrywiseExport'):
                                        Gen_CountrywiseExport_All = Gen_CountrywiseExport_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_localPurchTerms'):
                                        Gen_localPurchTerms_All = Gen_localPurchTerms_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'gen_ImportTerms'):
                                        gen_ImportTerms_All = gen_ImportTerms_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_Countrywise Purchase'):
                                        Gen_CountrywisePurchase_All = Gen_CountrywisePurchase_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'General_Source'):
                                        General_Source_All = General_Source_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'General_Telephone'):
                                        General_Telephone_All = General_Telephone_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Gen_SIC'):
                                        Gen_SIC_All = Gen_SIC_All.append(name)


                                    # HISTORY

                                    elif (name.at[0,  "SheetName"] == 'History Non loop'):
                                        History_Non_loop_All = History_Non_loop_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History Address'):
                                        History_Address_All = History_Address_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_Background'):
                                        History_Background_All = History_Background_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_auditor'):
                                        History_auditor_All = History_auditor_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_StockExchange'):
                                        History_StockExchange_All = History_StockExchange_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_Authorize'):
                                        History_Authorize_All = History_Authorize_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_Issued'):
                                        History_Issued_All = History_Issued_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_Paidup'):
                                        History_Paidup_All = History_Paidup_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'History_Sharehold'):
                                        History_Sharehold_All = History_Sharehold_All.append(name)

                                    # Mgmt Loop_Non Loop
                                    elif (name.at[0,  "SheetName"] == 'Mgmt Loop_Non Loop'):
                                        Mgmt_Loop_Non_Loop_All = Mgmt_Loop_Non_Loop_All.append(name)

                                    # Relcon

                                    elif (name.at[0,  "SheetName"] == 'Relcn Non loop'):
                                        Relcn_Non_loop_All = Relcn_Non_loop_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Rel_Subsidiary'):
                                        Rel_Subsidiary_All = Rel_Subsidiary_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Rel_Affiliate'):
                                        Rel_Affiliate_All = Rel_Affiliate_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Rel_Branch'):
                                        Rel_Branch_All = Rel_Branch_All.append(name)

                                    # Cinvest Loop, Bank Loop_NonLoop, Supplier looping

                                    elif (name.at[0,  "SheetName"] == 'Cinvest Loop'):
                                        Cinvest_Loop_All = Cinvest_Loop_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Bank Loop_Non Loop'):
                                        Bank_Loop_NonLoop_All = Bank_Loop_NonLoop_All.append(name)
                                    elif (name.at[0,  "SheetName"] == 'Supplier looping'):
                                        Supplier_looping_All = Supplier_looping_All.append(name)

                                    Status = "File Processed"
                                    # StatsList.append(
                                    #     [name.at[0,  'DUNS_General_aa'],  FileWithoutExtn,  name.at[0,  "SheetName"],  Status,
                                    #      str(today) + "_" + str(logtime),
                                    #      "Data Available - Insert Data"])
                                    StatsList.append(
                                        [name.at[0, 'duns_no'], FileWithoutExtn, name.at[0, "SheetName"],
                                         Status,str(today) + "_" + str(logtime),"Data Available - Insert Data"])
                                    if (df_FinNonFin.at[name.at[0,  "SheetName"],  "Deletion"] == "Y"):
                                        # df_FinNonFin.at[name.at[0,  "SheetName"],  "Looping/NonLooping"] == "L" and
                                        # s = pd.DataFrame(name.DUNS_General_aa.unique(), columns=['DUNS_General_aa'])
                                        s = pd.DataFrame(name.duns_no.unique(), columns=['duns_no'])
                                        if (name.at[0,  "SheetName"] == 'Gen_Insurance'):
                                            Gen_Insurance_All_Delete = Gen_Insurance_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_Award'):
                                            Gen_Award_All_Delete = Gen_Award_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_Membership'):
                                            Gen_Membership_All_Delete = Gen_Membership_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_Registration'):
                                            Gen_Registration_All_Delete = Gen_Registration_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_TradeStyle'):
                                            Gen_TradeStyle_All_Delete = Gen_TradeStyle_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_ProvFins'):
                                            Gen_ProvFins_All_Delete = Gen_ProvFins_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_ISOCert'):
                                            Gen_ISOCert_All_Delete = Gen_ISOCert_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_Product'):
                                            Gen_Product_All_Delete = Gen_Product_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_CustomerTable'):
                                            Gen_CustomerTable_All_Delete = Gen_CustomerTable_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_ExportTerms'):
                                            Gen_ExportTerms_All_Delete = Gen_ExportTerms_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_LocalSalesTerms'):
                                            Gen_LocalSalesTerms_All_Delete = Gen_LocalSalesTerms_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_CountrywiseExport'):
                                            Gen_CountrywiseExport_All_Delete = Gen_CountrywiseExport_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_localPurchTerms'):
                                            Gen_localPurchTerms_All_Delete = Gen_localPurchTerms_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'gen_ImportTerms'):
                                            gen_ImportTerms_All_Delete = gen_ImportTerms_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_Countrywise Purchase'):
                                            Gen_Countrywise_Purchase_All_Delete = Gen_Countrywise_Purchase_All_Delete.append(s)

                                        elif (name.at[0,  "SheetName"] == 'General_Telephone'):
                                            General_Telephone_All_Delete = General_Telephone_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Gen_SIC'):
                                            Gen_SIC_All_Delete = Gen_SIC_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History Address'):
                                            History_Address_All_Delete = History_Address_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_Background'):
                                            History_Background_All_Delete = History_Background_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_auditor'):
                                            History_auditor_All_Delete = History_auditor_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_StockExchange'):
                                            History_StockExchange_All_Delete = History_StockExchange_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_Authorize'):
                                            History_Authorize_All_Delete = History_Authorize_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_Issued'):
                                            History_Issued_All_Delete = History_Issued_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_Paidup'):
                                            History_Paidup_All_Delete = History_Paidup_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'History_Sharehold'):
                                            History_Sharehold_All_Delete = History_Sharehold_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Mgmt Loop_Non Loop'):
                                            Mgmt_Loop_Non_Loop_All_Delete = Mgmt_Loop_Non_Loop_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Rel_Subsidiary'):
                                            Rel_Subsidiary_All_Delete = Rel_Subsidiary_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Rel_Affiliate'):
                                            Rel_Affiliate_All_Delete = Rel_Affiliate_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Rel_Branch'):
                                            Rel_Branch_All_Delete = Rel_Branch_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Bank Loop_Non Loop'):
                                            Bank_Loop_Non_Loop_All_Delete = Bank_Loop_Non_Loop_All_Delete.append(s)
                                        elif (name.at[0,  "SheetName"] == 'Supplier looping'):
                                            Supplier_looping_All_Delete = Supplier_looping_All_Delete.append(s)
                                        elif (name.at[0, "SheetName"] == 'Cinvest Loop'):
                                            Cinvest_Loop_All_Delete = Cinvest_Loop_All_Delete.append(s)
                                        Status = "File Processed"

                                        # StatsList.append(
                                        #     [name.at[0,  'DUNS_General_aa'],  FileWithoutExtn,
                                        #      name.at[0,  "SheetName"],
                                        #      Status,
                                        #      str(today) + "_" + str(logtime),
                                        #      "Data Available - Delete Data"])
                                        StatsList.append(
                                            [name.at[0, 'duns_no'], FileWithoutExtn,
                                             name.at[0, "SheetName"],
                                             Status,
                                             str(today) + "_" + str(logtime),
                                             "Data Available - Delete Data"])

                                else:
                                    Status = "File Processed"
                                    StatsList.append(
                                        ["",  FileWithoutExtn,  str(key),  Status,
                                         str(today) + "_" + str(logtime),
                                         "DUNS Number validation failed"])
                            else:
                                logger.info("{} worksheet is empty".format(str(key)))
                                Status = "File Processed"
                                StatsList.append(
                                    ["",  FileWithoutExtn,  str(key),  Status,
                                     str(today) + "_" + str(logtime),
                                     "No Data Found"])

                        except Exception as ex:
                            Status = 'File Not Processed'
                            StatsList.append(
                                ["",  FileWithoutExtn,  str(key),  Status,  str(today) + "_" + str(logtime),
                                 str(ex)])
                            logger.error(filename)
                            os.remove( InPutFilesFinancial["InPutFilesFinancial"] + filename)
                            logger.error(str(ex))
                            continue

                    # Remove the Input file once processing is over
                    os.remove(os.path.join(Inputdirectory, file))

                except Exception as ex:
                    Status = 'File Not Processed'
                    StatsList.append(["",  FileWithoutExtn,  "",  Status,  str(today) + "_" + str(logtime),
                                      str(ex)])
                    logger.error(filename)
                    logger.error(str(ex))
                    continue
            # CSV Generation and Fields ID Insertion

            # -----------------------------------------------------------------------------------------------------------
            '''NON LOOPING CSV FILE GENERATION'''
            # -----------------------------------------------------------------------------------------------------------
            # Non Looping
            todaysDate = datetime.now().strftime("%d-%m-%y %H_%M_%S")
            if not Generalnonloop_All.empty:
                Generalnonloop_All=self.validate_data(Generalnonloop_All,
                                                      {'International_General:ev': 'Int64',
                      'International_General:dw': 'Int64',
                      'Total employees_General:cr': 'Int64',
                      'Local_General:et':'Int64',
                      'Local purchase terms_General:fe':'Int64',
                      'Import terms_General:fj':'Int64','Local_General:du':'Int64',
                      'Local sales terms_General:ee':"Int64",
                      'Export sales terms_General:ej':'Int64'},logger)
                Generalnonloop_All.to_csv(OutPutFilesReportToDews['OutPutFilesReportToDews'] + "Generalnonloop_All_"+todaysDate+"_"+TokenDict["General_non_loop_Insert"]+".csv",
                                    index=False)

            if not General_Source_All.empty:
                self.UpdateFieldIDDataframes(General_Source_All, logger)
                General_Source_All.to_csv(OutPutFilesReportToDews['OutPutFilesReportToDews'] + "General_Source_All_"+todaysDate+"_"+TokenDict["General_Source_Insert"]+".csv",
                                   index=False)
            if not History_Non_loop_All.empty:

                History_Non_loop_All['Shareholding as at_History:pc'] = pd.to_datetime(History_Non_loop_All['Shareholding as at_History:pc'], errors='coerce')
                History_Non_loop_All['Shareholding as at_History:pc'] = History_Non_loop_All['Shareholding as at_History:pc'].dt.strftime('%Y-%m-%d')

                History_Non_loop_All['Last AGM date_History:ju'] = pd.to_datetime(
                    History_Non_loop_All['Last AGM date_History:ju'], errors='coerce')
                History_Non_loop_All['Last AGM date_History:ju'] = History_Non_loop_All[
                    'Last AGM date_History:ju'].dt.strftime('%Y-%m-%d')

                History_Non_loop_All['Last financial end date_History:jw'] = pd.to_datetime(
                    History_Non_loop_All['Last financial end date_History:jw'], errors='coerce')
                History_Non_loop_All['Last financial end date_History:jw'] = History_Non_loop_All[
                    'Last financial end date_History:jw'].dt.strftime('%Y-%m-%d')

                History_Non_loop_All = History_Non_loop_All.astype \
                    ({'Ordinary Shares (number)_History:gw': 'Int64',
                      'Pref Shares (numbers)_History:ha': 'Int64',
                      'Shares publicly listed_History:gd':'Int64',
                      'Ordinary shares (Amount)_History:gt':'Int64',
                      'Par value_History:gu':'Int64'},
                     errors='ignore')

                History_Non_loop_All.to_csv(OutPutFilesReportToDews['OutPutFilesReportToDews'] + "History_Non_loop_All_"+todaysDate+"_"+TokenDict["History_Non_loop_Insert"]+".csv",
                                      index=False)
            if not Relcn_Non_loop_All.empty:
                Relcn_Non_loop_All = Relcn_Non_loop_All.astype \
                    ({'DUNS_Relcon:DE': 'Int64',
                      'Ultimate parent DUNS_Relcon:DZ':'Int64'},
                     errors='ignore')
                Relcn_Non_loop_All.to_csv(OutPutFilesReportToDews['OutPutFilesReportToDews'] + "Relcn_Non_loop_All_"+todaysDate+"_"+TokenDict["Relcn_Non_loop_Insert"]+".csv",
                                      index=False)

            # -----------------------------------------------------------------------------------------------------------
            '''LOOPING CSV FILE GENERATION'''
            # -----------------------------------------------------------------------------------------------------------
            # # Looping
            if not Gen_Insurance_All.empty:
                self.UpdateFieldIDDataframes(Gen_Insurance_All,  logger)
                Gen_Insurance_All.to_csv(Outputdirectory + "Gen_Insurance_All_"+todaysDate+"_"+TokenDict["Gen_Insurance_Insert"]+".csv",
                                         index=False)
            if not Gen_Award_All.empty:
                self.UpdateFieldIDDataframes(Gen_Award_All,  logger)
                Gen_Award_All.to_csv(Outputdirectory + "Gen_Award_All_"+todaysDate+"_"+TokenDict["Gen_Award_Insert"]+".csv",
                                         index=False)
            if not Gen_Membership_All.empty:
                self.UpdateFieldIDDataframes(Gen_Membership_All,  logger)
                Gen_Membership_All.to_csv(
                    Outputdirectory + "Gen_Membership_All_"+todaysDate+"_"+TokenDict["Gen_Membership_Insert"]+".csv",
                    index=False)
            if not Gen_Registration_All.empty:
                self.UpdateFieldIDDataframes(Gen_Registration_All,  logger)
                Gen_Registration_All.to_csv(
                    Outputdirectory + "Gen_Registration_All_"+todaysDate+"_"+TokenDict["Gen_Registration_Insert"]+".csv",
                    index=False)
            if not Gen_TradeStyle_All.empty:
                self.UpdateFieldIDDataframes(Gen_TradeStyle_All,  logger)
                Gen_TradeStyle_All.to_csv(Outputdirectory + "Gen_TradeStyle_All_"+todaysDate+"_"+TokenDict["Gen_TradeStyle_Insert"]+".csv",
                                         index=False)
            if not Gen_ProvFins_All.empty:
                self.UpdateFieldIDDataframes(Gen_ProvFins_All,  logger)
                Gen_ProvFins_All = self.validate_data(Gen_ProvFins_All,
                                                           {'Months covered_General:qo': 'Int64'},logger)
                Gen_ProvFins_All.to_csv(Outputdirectory + "Gen_ProvFins_All_"+todaysDate+"_"+TokenDict["Gen_ProvFins_Insert"]+".csv",
                                         index=False)
            if not Gen_ISOCert_All.empty:
                self.UpdateFieldIDDataframes(Gen_ISOCert_All,  logger)
                Gen_ISOCert_All.to_csv(
                    Outputdirectory + "Gen_ISOCert_All_"+todaysDate+"_"+TokenDict["Gen_ISOCert_Insert"]+".csv",
                    index=False)
            if not Gen_Product_All.empty:
                self.UpdateFieldIDDataframes(Gen_Product_All,  logger)
                Gen_Product_All.to_csv(
                    Outputdirectory + "Gen_Product_All_"+todaysDate+"_"+TokenDict["Gen_Product_Insert"]+".csv",
                    index=False)
            if not Gen_CustomerTable_All.empty:
                self.UpdateFieldIDDataframes(Gen_CustomerTable_All,  logger)
                Gen_CustomerTable_All = self.validate_data(Gen_CustomerTable_All, {'Length of relationship_General:PR': 'Int64'},logger)
                Gen_CustomerTable_All.to_csv(Outputdirectory + "Gen_CustomerTable_All_"+todaysDate+"_"+TokenDict["Gen_CustomerTable_Insert"]+".csv",
                                         index=False)
            if not Gen_ExportTerms_All.empty:
                self.UpdateFieldIDDataframes(Gen_ExportTerms_All,  logger)
                Gen_ExportTerms_All = self.validate_data(Gen_ExportTerms_All,
                                                           {'Net days_General:el': 'Int64',
                                                            'To net days_General:em':'Int64'}, logger)

                Gen_ExportTerms_All.to_csv(Outputdirectory + "Gen_ExportTerms_All_"+todaysDate+"_"+TokenDict["Gen_ExportTerms_Insert"]+".csv",
                                         index=False)
            if not Gen_LocalSalesTerms_All.empty:
                self.UpdateFieldIDDataframes(Gen_LocalSalesTerms_All,  logger)
                # Gen_LocalSalesTerms_All = Gen_LocalSalesTerms_All.astype({'general:eh': 'Int64','general:ei': 'Int64'}, errors='ignore')

                Gen_LocalSalesTerms_All = Gen_LocalSalesTerms_All.astype \
                    ({'Net days_General:eh': 'Int64',
                      'To net days_General:ei': 'Int64'},
                     errors='ignore')
                Gen_LocalSalesTerms_All.to_csv(
                    Outputdirectory + "Gen_LocalSalesTerms_All_"+todaysDate+"_"+TokenDict["Gen_LocalSalesTerms_Insert"]+".csv",
                    index=False)
            if not Gen_CountrywiseExport_All.empty:
                self.UpdateFieldIDDataframes(Gen_CountrywiseExport_All,  logger)
                Gen_CountrywiseExport_All.to_csv(
                    Outputdirectory + "Gen_CountrywiseExport_All_"+todaysDate+"_"+TokenDict["Gen_CountrywiseExport_Insert"]+".csv",
                    index=False)
            if not Gen_localPurchTerms_All.empty:
                self.UpdateFieldIDDataframes(Gen_localPurchTerms_All,  logger)
                # Gen_localPurchTerms_All = Gen_localPurchTerms_All.astype \
                #     ({'general:fh': 'Int64'},errors='ignore')
                Gen_localPurchTerms_All = Gen_localPurchTerms_All.astype \
                    ({'Net days_General:fh': 'Int64',
                      'To net days_General:fi': 'Int64'},
                     errors='ignore')
                Gen_localPurchTerms_All.to_csv(Outputdirectory + "Gen_localPurchTerms_All_"+todaysDate+"_"+TokenDict["Gen_localPurchTerms_Insert"]+".csv",
                                         index=False)
            if not gen_ImportTerms_All.empty:
                self.UpdateFieldIDDataframes(gen_ImportTerms_All,  logger)
                gen_ImportTerms_All = self.validate_data(gen_ImportTerms_All,
                                                         {'Net days_General:fl': 'Int64',
                                                          'To net days_General:fm': 'Int64'}, logger)
                # gen_ImportTerms_All = gen_ImportTerms_All.astype \
                #     ({'Net days_General:fl': 'Int64'},
                #      errors='ignore')
                gen_ImportTerms_All.to_csv(Outputdirectory + "gen_ImportTerms_All_"+todaysDate+"_"+TokenDict["gen_ImportTerms_Insert"]+".csv",
                                         index=False)
            if not Gen_CountrywisePurchase_All.empty:
                self.UpdateFieldIDDataframes(Gen_CountrywisePurchase_All,  logger)
                Gen_CountrywisePurchase_All.to_csv(
                    Outputdirectory + "Gen_CountrywisePurchase_All_"+todaysDate+"_"+TokenDict["Gen_Countrywise_Purchase_Insert"]+".csv",
                    index=False)
            if not General_Telephone_All.empty:
                self.UpdateFieldIDDataframes(General_Telephone_All,  logger)
                General_Telephone_All = self.validate_data(General_Telephone_All, {'Area code_General:ar':'Int64'},
                                                        logger)
                General_Telephone_All.to_csv(
                    Outputdirectory + "General_Telephone_All_"+todaysDate+"_"+TokenDict["General_Telephone_Insert"]+".csv",
                    index=False)
            if not Gen_SIC_All.empty:
                self.UpdateFieldIDDataframes(Gen_SIC_All,  logger)
                Gen_SIC_All.to_csv(
                    Outputdirectory + "Gen_SIC_All_"+todaysDate+"_"+TokenDict["Gen_SIC_Insert"]+".csv",
                    index=False)
            if not History_Address_All.empty:
                self.UpdateFieldIDDataframes(History_Address_All,  logger)
                History_Address_All.to_csv(
                    Outputdirectory + "History_Address_All_"+todaysDate+"_"+TokenDict["History_Address_Insert"]+".csv",
                    index=False)
            if not History_Background_All.empty:
                self.UpdateFieldIDDataframes(History_Background_All,  logger)
                History_Background_All.to_csv(
                    Outputdirectory + "History_Background_All_"+todaysDate+"_"+TokenDict["History_Background_Insert"]+".csv",
                    index=False)
            if not History_auditor_All.empty:
                self.UpdateFieldIDDataframes(History_auditor_All,  logger)
                History_auditor_All.to_csv(
                    Outputdirectory + "History_auditor_All_"+todaysDate+"_"+TokenDict["History_auditor_Insert"]+".csv",
                    index=False)
            if not History_StockExchange_All.empty:
                self.UpdateFieldIDDataframes(History_StockExchange_All,  logger)
                History_StockExchange_All.to_csv(
                    Outputdirectory + "History_StockExchange_All_"+todaysDate+"_"+TokenDict["History_StockExchange_Insert"]+".csv",
                    index=False)
            if not History_Authorize_All.empty:
                self.UpdateFieldIDDataframes(History_Authorize_All,  logger)
                History_Authorize_All.to_csv(
                    Outputdirectory + "History_Authorize_All_"+todaysDate+"_"+TokenDict["History_Authorize_Insert"]+".csv",
                    index=False)
            if not History_Issued_All.empty:
                self.UpdateFieldIDDataframes(History_Issued_All,  logger)
                History_Issued_All.to_csv(
                    Outputdirectory + "History_Issued_All_"+todaysDate+"_"+TokenDict["History_Issued_Insert"]+".csv",
                    index=False)
            if not History_Paidup_All.empty:
                self.UpdateFieldIDDataframes(History_Paidup_All,  logger)
                History_Paidup_All = self.validate_data(History_Paidup_All, {'Amount_History:hj': 'Int64'},logger)
                History_Paidup_All.to_csv(
                    Outputdirectory + "History_Paidup_All_"+todaysDate+"_"+TokenDict["History_Paidup_Insert"]+".csv",
                    index=False)
            if not History_Sharehold_All.empty:
                self.UpdateFieldIDDataframes(History_Sharehold_All,  logger)
                History_Sharehold_All = History_Sharehold_All.astype \
                    ({'Ord shares held(number)_History:ik': 'Int64',
                      'Pref shares held (number)_History:io': 'Int64'},
                     errors='ignore')

                # '% held_History:im': 'float16'
                # , float_format = '%.2f'
                History_Sharehold_All.to_csv(
                    Outputdirectory + "History_Sharehold_All_"+todaysDate+"_"+TokenDict["History_Sharehold_Insert"]+".csv",
                    index=False,float_format = '%.2f')
            if not Mgmt_Loop_Non_Loop_All.empty:
                self.UpdateFieldIDDataframes(Mgmt_Loop_Non_Loop_All,  logger)
                # Mgmt_Loop_Non_Loop_All = Mgmt_Loop_Non_Loop_All.astype \
                #     ({'CEO_Mgmtbac:EC': 'Int64'},errors='ignore')
                Mgmt_Loop_Non_Loop_All = self.validate_data(Mgmt_Loop_Non_Loop_All, {'CEO_Mgmtbac:EC':'Int64','Identity Number_Mgmtbac:EL':'Int64'},logger)
                Mgmt_Loop_Non_Loop_All['Identity Number_Mgmtbac:EL'].replace('0000<NA>','',inplace=True)
                Mgmt_Loop_Non_Loop_All.to_csv(
                    Outputdirectory + "Mgmt_Loop_Non_Loop_All_"+todaysDate+"_"+TokenDict["Mgmt_Loop_Non_Loop_Insert"]+".csv",
                    index=False)
            if not Rel_Subsidiary_All.empty:
                self.UpdateFieldIDDataframes(Rel_Subsidiary_All,  logger)
                Rel_Subsidiary_All.to_csv(
                    Outputdirectory + "Rel_Subsidiary_All_"+todaysDate+"_"+TokenDict["Rel_Subsidiary_Insert"]+".csv",
                    index=False)
            if not Rel_Affiliate_All.empty:

                self.UpdateFieldIDDataframes(Rel_Affiliate_All,  logger)

                Rel_Affiliate_All = self.validate_data(Rel_Affiliate_All,
                                                         {'Affiliate owns shares in subject_Relcon:BX': 'Int64',
                                                          'This affiliate is a joint venture_Recon:BF': 'Int64',
                                                          'Affiliate DUNS_Relcon:BG': 'Int64'}, logger)
                Rel_Affiliate_All.to_csv(
                    Outputdirectory + "Rel_Affiliate_All_"+todaysDate+"_"+TokenDict["Rel_Affiliate_Insert"]+".csv",
                    index=False)
            if not Rel_Branch_All.empty:
                self.UpdateFieldIDDataframes(Rel_Branch_All,  logger)
                Rel_Branch_All = self.validate_data(Rel_Branch_All, {'Post code_Relcon:yf': 'Int64'},logger)
                Rel_Branch_All.to_csv(
                    Outputdirectory + "Rel_Branch_All_"+todaysDate+"_"+TokenDict["Rel_Branch_Insert"]+".csv",
                    index=False)
            if not Bank_Loop_NonLoop_All.empty:
                self.UpdateFieldIDDataframes(Bank_Loop_NonLoop_All,  logger)
                # Post code_Bank: bh
                Bank_Loop_NonLoop_All = self.validate_data(Bank_Loop_NonLoop_All, {'Post code_Bank:bh': 'Int64'},logger)
                Bank_Loop_NonLoop_All.to_csv(
                    Outputdirectory + "Bank_Loop_NonLoop_All_"+todaysDate+"_"+TokenDict["Bank_Loop_Non_Loop_Insert"]+".csv",
                    index=False)
            if not Supplier_looping_All.empty:
                self.UpdateFieldIDDataframes(Supplier_looping_All,  logger)
                Supplier_looping_All['Report date_Supplier:la'] = pd.to_datetime(Supplier_looping_All['Report date_Supplier:la'],errors='coerce')
                Supplier_looping_All['Report date_Supplier:la'] = Supplier_looping_All['Report date_Supplier:la'].dt.strftime('%Y-%m-%d')
                Supplier_looping_All.to_csv(
                    Outputdirectory + "Supplier_looping_All_"+todaysDate+"_"+TokenDict["Supplier_looping_Insert"]+".csv",
                    index=False)
            if not Cinvest_Loop_All.empty:
                Cinvest_Loop_All.to_csv(OutPutFilesReportToDews['OutPutFilesReportToDews'] + "Cinvest_Loop_All_"+todaysDate+"_"+TokenDict["Cinvest_Loop_Insert"]+".csv",
                                      index=False)

            #----------------------------------------------------------------------------------------
            '''DELETE SECTION'''
            # ----------------------------------------------------------------------------------------
            # Deletion Looping
            if not Gen_Insurance_All_Delete.empty:
                # deleting the empty rows

                Gen_Insurance_All_Delete.dropna(how='all', inplace=True)
                Gen_Insurance_All_Delete.to_csv(Outputdirectory+"Gen_Insurance_All_Delete_"+todaysDate+"_"+TokenDict["Gen_Insurance_Delete"]+".csv")
            if not Gen_Award_All_Delete.empty:
                Gen_Award_All_Delete.dropna(how='all', inplace=True)
                Gen_Award_All_Delete.to_csv(Outputdirectory+"Gen_Award_All_Delete_"+todaysDate+"_"+TokenDict["Gen_Award_Delete"]+".csv")
            if not Gen_Membership_All_Delete.empty:
                Gen_Membership_All_Delete.dropna(how='all', inplace=True)
                Gen_Membership_All_Delete.to_csv(Outputdirectory + "Gen_Membership_All_Delete_"+todaysDate+"_"+TokenDict["Gen_Membership_Delete"]+".csv")
            if not Gen_Registration_All_Delete.empty:
                Gen_Registration_All_Delete.dropna(how='all', inplace=True)
                Gen_Registration_All_Delete.to_csv(Outputdirectory +"Gen_Registration_All_Delete_"+todaysDate+"_"+TokenDict["Gen_Registration_Delete"]+".csv")
            if not Gen_TradeStyle_All_Delete.empty:
                Gen_TradeStyle_All_Delete.dropna(how='all', inplace=True)
                Gen_TradeStyle_All_Delete.to_csv(Outputdirectory + "Gen_TradeStyle_All_Delete_"+todaysDate+"_"+TokenDict["Gen_TradeStyle_Delete"]+".csv")
            if not Gen_ProvFins_All_Delete.empty:
                Gen_ProvFins_All_Delete.dropna(how='all', inplace=True)
                Gen_ProvFins_All_Delete.to_csv(Outputdirectory + "Gen_ProvFins_All_Delete_"+todaysDate+"_"+TokenDict["Gen_ProvFins_Delete"]+".csv")
            if not Gen_ISOCert_All_Delete.empty:
                Gen_ISOCert_All_Delete.dropna(how='all', inplace=True)
                Gen_ISOCert_All_Delete.to_csv(Outputdirectory + "Gen_ISOCert_All_Delete_"+todaysDate+"_"+TokenDict["Gen_ISOCert_Delete"]+".csv")
            if not Gen_Product_All_Delete.empty:
                Gen_Product_All_Delete.dropna(how='all', inplace=True)
                Gen_Product_All_Delete.to_csv(Outputdirectory + "Gen_Product_All_Delete_"+todaysDate+"_"+TokenDict["Gen_Product_Delete"]+".csv")
            if not Gen_CustomerTable_All_Delete.empty:
                Gen_CustomerTable_All_Delete.dropna(how='all', inplace=True)
                Gen_CustomerTable_All_Delete.to_csv(Outputdirectory + "Gen_CustomerTable_All_Delete_"+todaysDate+"_"+TokenDict["Gen_CustomerTable_Delete"]+".csv")
            if not Gen_ExportTerms_All_Delete.empty:
                Gen_ExportTerms_All_Delete.dropna(how='all', inplace=True)
                Gen_ExportTerms_All_Delete.to_csv(Outputdirectory + "Gen_ExportTerms_All_Delete_"+todaysDate+"_"+TokenDict["Gen_ExportTerms_Delete"]+".csv")
            if not Gen_LocalSalesTerms_All_Delete.empty:
                Gen_LocalSalesTerms_All_Delete.dropna(how='all', inplace=True)
                Gen_LocalSalesTerms_All_Delete.to_csv(Outputdirectory + "Gen_LocalSalesTerms_All_Delete_"+todaysDate+"_"+TokenDict["Gen_LocalSalesTerms_Delete"]+".csv")
            if not Gen_CountrywiseExport_All_Delete.empty:
                Gen_CountrywiseExport_All_Delete.dropna(how='all', inplace=True)
                Gen_CountrywiseExport_All_Delete.to_csv(Outputdirectory + "Gen_CountrywiseExport_All_Delete_"+todaysDate+"_"+TokenDict["Gen_CountrywiseExport_Delete"]+".csv")
            if not Gen_localPurchTerms_All_Delete.empty:
                Gen_localPurchTerms_All_Delete.dropna(how='all', inplace=True)
                Gen_localPurchTerms_All_Delete.to_csv(Outputdirectory + "Gen_localPurchTerms_All_Delete_"+todaysDate+"_"+TokenDict["Gen_localPurchTerms_Delete"]+".csv")
            if not gen_ImportTerms_All_Delete.empty:
                gen_ImportTerms_All_Delete.dropna(how='all', inplace=True)
                gen_ImportTerms_All_Delete.to_csv(Outputdirectory + "gen_ImportTerms_All_Delete_"+todaysDate+"_"+TokenDict["gen_ImportTerms_Delete"]+".csv")
            if not Gen_Countrywise_Purchase_All_Delete.empty:
                Gen_Countrywise_Purchase_All_Delete.dropna(how='all', inplace=True)
                Gen_Countrywise_Purchase_All_Delete.to_csv(Outputdirectory + "Gen_Countrywise_Purchase_All_Delete_"+todaysDate+"_"+TokenDict["Gen_Countrywise_Purchase_Delete"]+".csv")
            if not General_Telephone_All_Delete.empty:
                General_Telephone_All_Delete.dropna(how='all', inplace=True)
                General_Telephone_All_Delete.to_csv(Outputdirectory + "General_Telephone_All_Delete_"+todaysDate+"_"+TokenDict["General_Telephone_Delete"]+".csv")
            if not Gen_SIC_All_Delete.empty:
                Gen_SIC_All_Delete.dropna(how='all', inplace=True)
                Gen_SIC_All_Delete.to_csv(Outputdirectory + "Gen_SIC_All_Delete_"+todaysDate+"_"+TokenDict["Gen_SIC_Delete"]+".csv")
            if not History_Address_All_Delete.empty:
                History_Address_All_Delete.dropna(how='all', inplace=True)
                History_Address_All_Delete.to_csv(Outputdirectory + "History_Address_All_Delete_"+todaysDate+"_"+TokenDict["History_Address_Delete"]+".csv")
            if not History_Background_All_Delete.empty:
                History_Background_All_Delete.dropna(how='all', inplace=True)
                History_Background_All_Delete.to_csv(Outputdirectory + "History_Background_All_Delete_"+todaysDate+"_"+TokenDict["History_Background_Delete"]+".csv")
            if not History_auditor_All_Delete.empty:
                History_auditor_All_Delete.dropna(how='all', inplace=True)
                History_auditor_All_Delete.to_csv(Outputdirectory + "History_auditor_All_Delete_"+todaysDate+"_"+TokenDict["History_auditor_Delete"]+".csv")
            if not History_StockExchange_All_Delete.empty:
                History_StockExchange_All_Delete.dropna(how='all', inplace=True)
                History_StockExchange_All_Delete.to_csv(Outputdirectory + "History_StockExchange_All_Delete_"+todaysDate+"_"+TokenDict["History_StockExchange_Delete"]+".csv")
            if not History_Authorize_All_Delete.empty:
                History_Authorize_All_Delete.dropna(how='all', inplace=True)
                History_Authorize_All_Delete.to_csv(Outputdirectory + "History_Authorize_All_Delete_"+todaysDate+"_"+TokenDict["History_Authorize_Delete"]+".csv")
            if not History_Issued_All_Delete.empty:
                History_Issued_All_Delete.dropna(how='all', inplace=True)
                History_Issued_All_Delete.to_csv(Outputdirectory + "History_Issued_All_Delete_"+todaysDate+"_"+TokenDict["History_Issued_Delete"]+".csv")
            if not History_Paidup_All_Delete.empty:
                History_Paidup_All_Delete.dropna(how='all', inplace=True)
                History_Paidup_All_Delete.to_csv(Outputdirectory + "History_Paidup_All_Delete_"+todaysDate+"_"+TokenDict["History_Paidup_Delete"]+".csv")
            if not History_Sharehold_All_Delete.empty:
                History_Sharehold_All_Delete.dropna(how='all', inplace=True)
                History_Sharehold_All_Delete.to_csv(Outputdirectory + "History_Sharehold_All_Delete_"+todaysDate+"_"+TokenDict["History_Sharehold_Delete"]+".csv")
            if not Mgmt_Loop_Non_Loop_All_Delete.empty:
                Mgmt_Loop_Non_Loop_All_Delete.dropna(how='all', inplace=True)
                Mgmt_Loop_Non_Loop_All_Delete.to_csv(Outputdirectory + "Mgmt_Loop_Non_Loop_All_Delete_"+todaysDate+"_"+TokenDict["Mgmt_Loop_Non_Loop_Delete"]+".csv")
            if not Rel_Subsidiary_All_Delete.empty:
                Rel_Subsidiary_All_Delete.dropna(how='all', inplace=True)
                Rel_Subsidiary_All_Delete.to_csv(Outputdirectory + "Rel_Subsidiary_All_Delete_"+todaysDate+"_"+TokenDict["Rel_Subsidiary_Delete"]+".csv")
            if not Rel_Affiliate_All_Delete.empty:
                Rel_Affiliate_All_Delete.dropna(how='all', inplace=True)
                Rel_Affiliate_All_Delete.to_csv(Outputdirectory + "Rel_Affiliate_All_Delete_"+todaysDate+"_"+TokenDict["Rel_Affiliate_Delete"]+".csv")
            if not Rel_Branch_All_Delete.empty:
                Rel_Branch_All_Delete.dropna(how='all', inplace=True)
                Rel_Branch_All_Delete.to_csv(Outputdirectory + "Rel_Branch_All_Delete_"+todaysDate+"_"+TokenDict["Rel_Branch_Delete"]+".csv")
            if not Bank_Loop_Non_Loop_All_Delete.empty:
                Bank_Loop_Non_Loop_All_Delete.dropna(how='all', inplace=True)
                Bank_Loop_Non_Loop_All_Delete.to_csv(Outputdirectory + "Bank_Loop_Non_Loop_All_Delete_"+todaysDate+"_"+TokenDict["Bank_Loop_Non_Loop_Delete"]+".csv")
            if not Supplier_looping_All_Delete.empty:
                Supplier_looping_All_Delete.dropna(how='all', inplace=True)
                Supplier_looping_All_Delete.to_csv(Outputdirectory + "Supplier_looping_All_Delete_"+todaysDate+"_"+TokenDict["Supplier_looping_Delete"]+".csv")
            if not Cinvest_Loop_All_Delete.empty:
                Cinvest_Loop_All_Delete.dropna(how='all', inplace=True)
                Cinvest_Loop_All_Delete.to_csv(Outputdirectory + "Cinvest_Loop_All_Delete_"+todaysDate+"_"+TokenDict["Cinvest_Loop_Delete"]+".csv")


            # Upload CSV Files to Sharepoint
            MessageUploadOutput = self.sharepointoperation("Upload File", OutPutFilesReportToDews['OutPutFilesReportToDews'] + "",
                 logger, "CSV")
            if (MessageUploadOutput == "File uploaded"):
                logger.info("CSV files have been uploaded to sharepoint")

            # Upload files to FTP
            cnopts = pysftp.CnOpts()
            cnopts.hostkeys = paramiko.hostkeys.HostKeys(sftpauth["sftphostkey"])
            hostName=ftp["hostName"]
            userName = ftp["userName"]
            pswd =ftp["pswd"]

            try:
                with pysftp.Connection(host=hostName,  username=userName,
                                       password=pswd,  cnopts=cnopts) as sftp:
                    try:
                        logger.info("Connection established ... ")

                        with sftp.cd("/puts/"):
                            # print(sftp.listdir())
                            try:
                                # # Use put method to upload a file
                                # First trasfer all delete files

                                list_of_files = sorted(
                                    filter(lambda x: os.path.isfile(os.path.join(Outputdirectory, x)),
                                           os.listdir(Outputdirectory)))

                                for file in list_of_files:
                                    try:
                                        if not file.endswith(".csv"):
                                            continue
                                        if 'Delete' in file:
                                            sftp.put(os.path.join(Outputdirectory,  file),  confirm=False)
                                            # ,  confirm = False
                                            time.sleep(1)
                                            FTPStats.append(
                                                [file,  "FTP Transfer Successfull",  str(today) + "_" + str(logtime),
                                                 "No Remarks"])
                                            

                                            os.remove(os.path.join(Outputdirectory,  file))
                                    except Exception as ex:
                                        logger.error(str(ex))
                                        FTPStats.append(
                                            [file,  "FTP Transfer Failed",  str(today) + "_" + str(logtime),
                                             str(ex)])
                                        continue
                                time.sleep(1)
                                # second trasfer all Insert files

                                list_of_files = sorted(
                                    filter(lambda x: os.path.isfile(os.path.join(Outputdirectory, x)),
                                           os.listdir(Outputdirectory)))

                                for file in list_of_files:
                                    try:
                                        if not file.endswith(".csv"):
                                            continue
                                        if 'Delete' not in file:
                                            sftp.put(os.path.join(Outputdirectory,  file),  confirm=False)
                                            # ,  confirm = False
                                            time.sleep(1)
                                            FTPStats.append(
                                                [file,  "FTP Transfer Successfull",  str(today) + "_" + str(logtime),
                                                 "No Remarks"])

                                            os.remove(os.path.join(Outputdirectory,  file))
                                    except Exception as ex:
                                        logger.error(str(ex))
                                        logger.info ("delete file sent ")
                                        FTPStats.append(
                                            [file,  "FTP Transfer Failed",  str(today) + "_" + str(logtime),
                                             str(ex)])
                                        continue

                            except Exception as ex:
                                logger.error(str(ex))
                    except Exception as ex:

                        logger.error(str(ex))
                        FTPStats.append(
                            [file,  "FTP Transfer Failed",  str(today) + "_" + str(logtime),
                             str(ex)])

            except Exception as ex:
                logger.error("Connection Not established ... ")
                logger.error(str(ex))
                FTPStats.append(
                    ["Connection Not established ...",  "FTP Transfer Failed",  str(today) + "_" + str(logtime),
                     str(ex)])
            # Load stats file

            self.CreateStats(logger,  AutoFinancialObj,  StatsList, FTPStats, ValidationStatsList)

        except Exception as ex:
            logger.error(str(ex))

    # Create the Stats of the Process
    def CreateStats(self, logger, AutoFinancialObj, StatsList, FTPList,  ValidationStatsList):

        global Stats, ValidationStats, FTPStatdf, Validation_Stats_Current
        # global FTPStatdf
        FTPStatdf = pd.DataFrame()
        ValidationStats = pd.DataFrame()
        Validation_Stats_Current=pd.DataFrame()
        Stats = pd.DataFrame()

        Outputdirectory = OutPutFilesReportToDews["OutPutFilesReportToDews"]
        try:
            book = load_workbook(Outputdirectory + 'AutoReportToDews_Stats.xlsx')

            # Maintaining the Sharepoint folder stats
            DFSharepointFolderStatus = pd.concat([AutoFinancialObj.DFSharepointFolderDetails])
            DFSharepointFolderStatus.columns = ['Folder Name',  'File Count',  'Folder Status',  'ProcessedDateTime',
                                                'File download Count']

            # Maintaining the Worskheets Stats
            Stats = Stats.append(pd.DataFrame(StatsList,  columns=['DUNS Number',  'File Name',  'WorkSheet Name',  'Status',
                                                                  'Processed Date',  'Remarks']),  ignore_index=True)

            # Maintaining the FTP Stats
            FTPStatdf = FTPStatdf.append(
                pd.DataFrame(FTPList,  columns=['File Name',  'Status',  'Processed Date',  'Remarks']),  ignore_index=True)

            # Maintaining the Validation Stats
            ValidationStats = ValidationStats.append(pd.DataFrame(ValidationStatsList,  columns=['DUNS Number',  'File Name',  'File Type',  'Status',
                                                                  'Processed Date',  'Remarks']),  ignore_index=True)

            # Creating writer Object
            writer = pd.ExcelWriter(Outputdirectory + 'AutoReportToDews_Stats.xlsx',
                                    engine='openpyxl')
            writer.book = book
            # writer.sheets = {ws.title: ws for ws in book.worksheets}

            writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))
            with writer:
                Stats.to_excel(writer,  sheet_name='CSV File log',
                               startrow=writer.sheets['CSV File log'].max_row,
                               index=False,  header=False)
                ValidationStats.to_excel(writer,  sheet_name='Intermediaryfile log',
                               startrow=writer.sheets['Intermediaryfile log'].max_row,
                               index=False,  header=False)
                FTPStatdf.to_excel(writer,  sheet_name='FTP Stats',
                                   startrow=writer.sheets['FTP Stats'].max_row,
                                   index=False,  header=False)
                DFSharepointFolderStatus.to_excel(writer,  sheet_name='SharePoint Folder Stats',
                                                  startrow=writer.sheets['SharePoint Folder Stats'].max_row,
                                                  index=False,  header=False)

            book.save(Outputdirectory + 'AutoReportToDews_Stats.xlsx')
            book.save(Outputdirectory + 'AutoReportToDews_Stats_' + str(today) + '.xlsx')

            MessageUploadOutput = self.sharepointoperation("Upload File", Outputdirectory + 'AutoReportToDews_Stats.xlsx',
                                                           logger,  "LOGS")
            if (MessageUploadOutput == "File uploaded"):
                logger.info("Output file has been uploaded to sharepoint")
                os.remove(os.path.join(Outputdirectory,
                                       'AutoReportToDews_Stats_' + str(today) + '.xlsx'))

        except Exception as ex:
            logger.error("Stats Error")
            logger.error(str(ex))

def main():
    # Defining Variables
    global logtime,  today

    # Object creation for Logging module and Email class
    loggerObj = Logs()
    mailobj = mail()

    # Define time to set in the log file name
    logtime = str(datetime.now().strftime('%H_%M_%S'))
    today = date.today().strftime('%d_%m_%Y')

    # define the file name
    logfilename = "ReportToDewsLogs_" + str(today) + ".txt"
    LogStatsPath = settings_ReportToDews.get('team_site_url') + '/Shared%20Documents/CBIG/Logs_Stats/'
    LogPath = settings_ReportToDews.get('team_site_url') + '/Shared%20Documents/CBIG/Logs_Stats/'
    logger = loggerObj.setup_logger('ReportToDewsLogs_',  LogFileReportToDews["LogFileReportToDews"] + logfilename)
    try:

        logger.info("*********************************START*****************************************************")
        Obj1 = AutoReportToDews()
        ProcessStartTime = time.time()
        Obj1.autofinnonfin(logger)
        ProcessEndTime = time.time()
        TotalElapsedTime = time.strftime("%H:%M:%S %Z",  time.gmtime(ProcessEndTime - ProcessStartTime))

        # Stats for the email => Total file ran, Failed and Success count
        if not Validation_Stats_Current.empty:
            DataFound = [0]
            DataNotFound = [0]
            df = Validation_Stats_Current["Status"].value_counts()
            try:
                DataFound = [df.at['Validation Successfull']]
            except Exception as ex:
                if str(ex) == "File Not Processed":
                    DataFound = [0]
            try:
                DataNotFound = [df.at['Validation Failed']]
            except Exception as ex:
                if str(ex.args).strip() == '(\'File Not Processed\', )':
                    DataNotFound = [0]
            try:
                details = {
                    "Total Files Run": [Validation_Stats_Current.shape[0]],
                    "Success Files": DataFound,
                    "Failed Files": DataNotFound,
                    "Processed DateTime": [date.today().strftime('%d_%m_%Y')],
                    "Run By": 'Auto Trigger Process',
                    "Total Execution Time": str(TotalElapsedTime)
                }
            except Exception as ex:
                details = {
                    "Total Files Run": [Validation_Stats_Current.shape[0]],
                    "Success Files": DataFound,
                    "Failed Files": DataNotFound,
                    "Processed DateTime": [date.today().strftime('%d_%m_%Y')],
                    "Run By": 'Auto Trigger Process',
                    "Total Execution Time": str(TotalElapsedTime)
                }

            df_stats = pd.DataFrame(details)
            output = build_table(df_stats,  'orange_light',  font_size='15px',  text_align='center',  width='auto', font_family='Open Sans, sans-serif')
        # send email only if files have failed in conversion
            if (DataNotFound != [0]):
                mailobj.SendEmailToStakeHolders("Report to Dews Upload Status-Failed ",  "DNBSystemMailDoNotReply@dnb.com",
                                                "khannapo@dnb.com",  output,  "Report to Dews Upload",  LogStatsPath)

        # Send email if FTP Transfer failes
        if not FTPStatdf.empty:
            DataFound = [0]
            DataNotFound = [0]
            df = FTPStatdf["Status"].value_counts()
            try:
                DataFound = [df.at['FTP Transfer Successfull']]
            except KeyError as ex:
                logger.error("KeyError"+str(ex))
                DataFound = [0]
            except Exception as ex:
                logger.error(str(ex))
                DataFound = [0]

            try:
                DataNotFound = [df.at['FTP Transfer Failed']]
            except Exception as ex:
                if str(ex.args).strip() == '(\'FTP Transfer Failed\', )':
                    DataNotFound = [0]

            try:
                detailsFTP = {
                    "Total Files Run": [FTPStatdf.shape[0]],
                    "Success Files": DataFound,
                    "Failed Files": DataNotFound,
                    "Processed DateTime": [date.today().strftime('%d_%m_%Y')],
                    "Run By": 'Auto Trigger Process',
                    "Total Execution Time": str(TotalElapsedTime)
                }
            except Exception as ex:
                detailsFTP = {
                    "Total Files Run": [FTPStatdf.shape[0]],
                    "Success Files": DataFound,
                    "Failed Files": DataNotFound,
                    "Processed DateTime": [date.today().strftime('%d_%m_%Y')],
                    "Run By": 'Auto Trigger Process',
                    "Total Execution Time": str(TotalElapsedTime)
                }

            df_ftpstats = pd.DataFrame(detailsFTP)
            output = build_table(df_ftpstats,  'orange_light',  font_size='15px',  text_align='center',  width='auto',
                                 font_family='Open Sans, sans-serif')
            # send email only if files have failed in conversion
            if (DataNotFound != [0]):
                mailobj.SendEmailToStakeHolders("Report to Dews Upload Status-Failed FTP Transfer",
                                                "DNBSystemMailDoNotReply@dnb.com",
                                                "khannapo@dnb.com",  output,  "Report to Dews - FTP Transfer Failed",  LogStatsPath)

        logger.info(f"{TotalElapsedTime} seconds to process and upload all Report to Dews files")

        MessageUploadLog = Obj1.sharepointoperation("Upload File",
                                                    LogFileReportToDews["LogFileReportToDews"] + logfilename,
                                                    logger,  "LOGS")

        if (MessageUploadLog == "File uploaded"):
            logger.info("log file has been uploaded to sharepoint")

        logger.info("**********************************END****************************************************")

    except Exception as ex:
        logger.error(str(ex))
        mailobj.SendEmailToStakeHolders("Report to Dews upload ERROR ",  "DNBSystemMailDoNotReply@dnb.com",
                                        "khannapo@dnb.com",  str(ex),  "Report to Dews Upload",  LogPath)

if __name__ == '__main__':
    main()
    quit()